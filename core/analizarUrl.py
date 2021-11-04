from openpyxl import Workbook
from openpyxl import load_workbook
from hashlib import md5
from virus_total_apis import PublicApi
import time
import socket


def analizarUrl(txt, excel, apikey):
    fo = open(txt, "r")
    libro = Workbook()
    hoja = libro.active
    hoja.cell(1, 1, "Url")
    hoja.cell(1, 2, "Fecha de análisis")
    hoja.cell(1, 3, "Total de análisis")
    hoja.cell(1, 4, "Análisis positivos")
    hoja.cell(1, 5, "Clasificación")
    api_key = apikey
    api = PublicApi(api_key)
    protocolo = socket.IPPROTO_TCP
    i = 2
    foo = open("socket.txt", "w")
    for line in fo:
        response = api.get_url_report(line)
        hoja.cell(i, 1, line)
        if response["response_code"] == 200:
            try:
                infoWeb = socket.getaddrinfo(line, 80, proto=protocolo)
                foo.write(line)
                foo.write("\n")
                foo.write(str(infoWeb[0]))
                foo.write("\n\n")
            except:
                pass
            hoja.cell(i, 2, response["results"]["scan_date"])
            hoja.cell(i, 3, len(response["results"]["scans"]))
            hoja.cell(i, 4, response["results"]["positives"])
            if response["results"]["positives"] <= 3:
                hoja.cell(i, 5, "Baja")
            elif (response["results"]["positives"] > 3 and response["results"]["positives"] <= 10):
                hoja.cell(i, 5, "Medio")
            elif response["results"]["positives"] > 10:
                hoja.cell(i, 5, "Alto")
        else:
            hoja.cell(i, 2, "Error de conexión")
        i += 1
        time.sleep(15)
    fo.close()
    foo.close()
    libro.save(excel)
