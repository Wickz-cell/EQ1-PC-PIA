from openpyxl import Workbook
from openpyxl import load_workbook
from hashlib import md5
from virus_total_apis import PublicApi
import time
import socket


def analizarUrl(txt, excel, apikey):
    print("1")
    fo = open(txt, "r")
    libro = Workbook()
    hoja = libro.active
    print("2")
    hoja.cell(1, 1, "Url")
    hoja.cell(1, 2, "Fecha de análisis")
    hoja.cell(1, 3, "Total de análisis")
    hoja.cell(1, 4, "Análisis positivos")
    hoja.cell(1, 5, "Clasificación")
    print("3")
    api_key = apikey
    api = PublicApi(api_key)
    protocolo = socket.IPPROTO_TCP
    i = 2
    foo = open("socket.txt", "w")
    print("4")
    for line in fo:
        response = api.get_url_report(line)
        hoja.cell(i, 1, line)
        print("5")
        if response["response_code"] == 200:
            try:
                infoWeb = socket.getaddrinfo(line, 80, proto=protocolo)
                print(infoWeb)
                print(infoWeb.type())
                foo.write(str(infoWeb))
                foo.write("\n")
            except:
                pass
            hoja.cell(i, 2, response["results"]["scan_date"])
            hoja.cell(i, 3, len(response["results"]["scans"]))
            hoja.cell(i, 4, response["results"]["positives"])
            print("6")
            if response["results"]["positives"] <= 3:
                hoja.cell(i, 5, "Baja")
                print("7")
            elif (response["results"]["positives"] > 3 and response["results"]["positives"] <= 10):
                hoja.cell(i, 5, "Medio")
                print("7")
            elif response["results"]["positives"] > 10:
                hoja.cell(i, 5, "Alto")
                print("7")
        else:
            hoja.cell(i, 2, "Error de conexión")
            print("8")
        i += 1
        time.sleep(15)
        print("9")
    fo.close()
    foo.close()
    print("10")
    libro.save(excel)

if "__name__" == "__main__":
    analizarUrl("urls_sospechosas.txt", "reporte.xlsx", "d08e178652d9db28ead746bee4495e0fd4fe91582d17b08e019320d92268f560")
